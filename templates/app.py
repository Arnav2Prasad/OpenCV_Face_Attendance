import cv2
import os
from flask import Flask, request, render_template
from datetime import date
from datetime import datetime
import numpy as np
from sklearn.neighbors import KNeighborsClassifier
import pandas as pd
import joblib
from openpyxl import Workbook, load_workbook


app = Flask(__name__)

nimgs = 10

imgBackground = cv2.imread("bg.png")

datetoday = date.today().strftime("%m_%d_%y")
datetoday2 = date.today().strftime("%d-%B-%Y")

face_detector = cv2.CascadeClassifier('haarcascade_frontalface_default.xml')

if not os.path.isdir('Attendance'):
    os.makedirs('Attendance')
if not os.path.isdir('static'):
    os.makedirs('static')
if not os.path.isdir('static/faces'):
    os.makedirs('static/faces')
if f'Attendance-{datetoday}.csv' not in os.listdir('Attendance'):
    with open(f'Attendance/Attendance-{datetoday}.csv', 'w') as f:
        f.write('Name,Roll,Time')

if 'students.xlsx' not in os.listdir():
    wb = Workbook()
    ws = wb.active
    ws.append(['Name', 'MIS'])
    wb.save('students.xlsx')

app_running = True

def totalreg():
    return len(os.listdir('static/faces'))

def extract_faces(img):
    try:
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        face_points = face_detector.detectMultiScale(gray, 1.2, 5, minSize=(20, 20))
        return face_points
    except:
        return []

def identify_face(facearray):
    model = joblib.load('static/face_recognition_model.pkl')
    return model.predict(facearray)

def train_model():
    faces = []
    labels = []
    userlist = os.listdir('static/faces')
    for user in userlist:
        for imgname in os.listdir(f'static/faces/{user}'):
            img = cv2.imread(f'static/faces/{user}/{imgname}')
            resized_face = cv2.resize(img, (50, 50))
            faces.append(resized_face.ravel())
            labels.append(user)
    faces = np.array(faces)
    knn = KNeighborsClassifier(n_neighbors=5)
    knn.fit(faces, labels)
    joblib.dump(knn, 'static/face_recognition_model.pkl')

def extract_attendance():
    df = pd.read_csv(f'Attendance/Attendance-{datetoday}.csv')
    names = df['Name']
    rolls = df['Roll']
    times = df['Time']
    l = len(df)
    return names, rolls, times, l

def add_attendance(name):
    username = name.split('_')[0]
    userid = name.split('_')[1]
    current_time = datetime.now().strftime("%H:%M:%S")

    if userid.strip():  # Check if userid is not an empty string
        try:
            userid_int = int(userid)
            df = pd.read_csv(f'Attendance/Attendance-{datetoday}.csv')
            if userid_int not in df['Roll'].values:
                with open(f'Attendance/Attendance-{datetoday}.csv', 'a') as f:
                    f.write(f'\n{username},{userid_int},{current_time}')
        except ValueError:
            print(f"Invalid userid: {userid}")
    else:
        print("Empty userid provided")

def update_students_excel(name, mis):
    wb = load_workbook('students.xlsx')
    ws = wb.active
    ws.append([name, mis])
    wb.save('students.xlsx')

def getallusers():
    userlist = os.listdir('static/faces')
    names = []
    rolls = []
    l = len(userlist)

    for i in userlist:
        name, roll = i.split('_')
        names.append(name)
        rolls.append(roll)

    return userlist, names, rolls, l

@app.route('/')
def new_home():
    names, rolls, times, l = extract_attendance()
    return render_template('new_home.html', names=names, rolls=rolls, times=times, l=l, totalreg=totalreg(), datetoday2=datetoday2)

@app.route('/attendance_page')
def display_attendance_page():
    names, rolls, times, l = extract_attendance()
    return render_template('attendance.html', names=names, rolls=rolls, times=times, l=l, totalreg=totalreg(), datetoday2=datetoday2)

@app.route('/newuser_page')
def display_newuser_page():
    names, rolls, times, l = extract_attendance()
    return render_template('newuser.html', names=names, rolls=rolls, times=times, l=l, totalreg=totalreg(), datetoday2=datetoday2)

@app.route('/qr_page')
def display_qr_page():
    names, rolls, times, l = extract_attendance()
    return render_template('qr.html', names=names, rolls=rolls, times=times, l=l, totalreg=totalreg(), datetoday2=datetoday2)

@app.route('/attendance', methods=['GET'])
def attendance():
    global app_running

    names, rolls, times, l = extract_attendance()

    if 'face_recognition_model.pkl' not in os.listdir('static'):
        return render_template('attendance.html', names=names, rolls=rolls, times=times, l=l, totalreg=totalreg(), datetoday2=datetoday2, mess='There is no trained model in the static folder. Please add a new face to continue.')

    ret = True
    cap = cv2.VideoCapture(0)
    while ret and app_running:
        ret, frame = cap.read()

        if len(extract_faces(frame)) > 0:
            (x, y, w, h) = extract_faces(frame)[0]
            cv2.rectangle(frame, (x, y), (x + w, y + h), (86, 32, 251), 1)
            cv2.rectangle(frame, (x, y), (x + w, y - 40), (86, 32, 251), -1)
            face = cv2.resize(frame[y:y + h, x:x + w], (50, 50))
            identified_person = identify_face(face.reshape(1, -1))[0]
            add_attendance(identified_person)
            cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 0, 255), 1)
            cv2.rectangle(frame, (x, y), (x + w, y + h), (50, 50, 255), 2)
            cv2.rectangle(frame, (x, y - 40), (x + w, y), (50, 50, 255), -1)
            cv2.putText(frame, f'{identified_person}', (x, y - 15), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 1)
            cv2.rectangle(frame, (x, y), (x + w, y + h), (50, 50, 255), 1)
        
        # Calculate the center position for the webcam window
        center_x = (imgBackground.shape[1] - frame.shape[1]) // 2
        center_y = (imgBackground.shape[0] - frame.shape[0]) // 2
        imgBackground[center_y:center_y + frame.shape[0], center_x:center_x + frame.shape[1]] = frame

        cv2.imshow('Attendance', imgBackground)
        if cv2.waitKey(1) == 27 or not app_running:
            app_running = False
            break

    cap.release()
    cv2.destroyAllWindows()
    names, rolls, times, l = extract_attendance()
    return render_template('attendance.html', names=names, rolls=rolls, times=times, l=l, totalreg=totalreg(), datetoday2=datetoday2)


@app.route('/newuser', methods=['GET', 'POST'])
def newuser():
    if request.method == 'POST':
        username = request.form['username']
        userid = request.form['userid']
        update_students_excel(username, userid)

        userimagefolder = 'static/faces/' + username + '_' + str(userid)
        if not os.path.isdir(userimagefolder):
            os.makedirs(userimagefolder)
        i, j = 0, 0
        cap = cv2.VideoCapture(0)
        while 1:
            _, frame = cap.read()
            faces = extract_faces(frame)
            for (x, y, w, h) in faces:
                cv2.rectangle(frame, (x, y), (x + w, y + h), (255, 0, 20), 2)
                cv2.putText(frame, f'Images Captured: {i}/{nimgs}', (30, 30),
                            cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 0, 20), 2, cv2.LINE_AA)
                if j % 5 == 0:
                    name = username + '_' + str(i) + '.jpg'
                    cv2.imwrite(userimagefolder + '/' + name, frame[y:y + h, x:x + w])
                    i += 1
                j += 1
            if j == nimgs * 5:
                break
            cv2.imshow('Adding new User', frame)
            if cv2.waitKey(1) == 27:
                break
        cap.release()
        cv2.destroyAllWindows()
        print('Training Model')
        train_model()
        names, rolls, times, l = extract_attendance()
        return render_template('newuser.html', names=names, rolls=rolls, times=times, l=l, totalreg=totalreg(), datetoday2=datetoday2)
    else:
        names, rolls, times, l = extract_attendance()
        return render_template('newuser.html', names=names, rolls=rolls, times=times, l=l, totalreg=totalreg(), datetoday2=datetoday2)

excel_file_path = 'students.xlsx'

# Read the Excel file into a DataFrame
df_login = pd.read_excel(excel_file_path)

# Print the DataFrame
df_login.rename(columns={'Name':'NAME'},inplace=True)
print(df_login)

import pandas as pd
import mysql.connector

# Replace these values with your MySQL connection details
host = 'localhost'
user = 'arnav-rppoop1'
password = 'Guitar@123'
database = 'db_1'

# Connect to MySQL
conn = mysql.connector.connect(host=host, user=user, password=password, database=database)

query = "SELECT * FROM result"



cursor = conn.cursor()
# cursor.execute(query)
# Fetch data into a DataFrame
df_sql = pd.read_sql(query, conn)

print(df_sql)


# Iterate over the rows using iloc
for i in range(len(df_login)):
    row = df_login.iloc[i]
    print("Row", i, ":", row['NAME'], row['MIS'])
    # Row index and attribute to check
    row_index = i  # Index of the row to check
    attribute_to_check = 'NAME'  # Attribute to check in df1

    # Get the value of the attribute to check
    value_to_check = df_login.loc[row_index, attribute_to_check]
    # Check if the row is present in df2
    # Check if the value is present in df2
    value_present_in_df2 = value_to_check in df_sql[attribute_to_check].values

    if not(value_present_in_df2):
        print(f"The value {value_to_check} from attribute {attribute_to_check} is not present in df2")
        print(type(df_login.loc[i,'NAME']))
        print(type(df_login.loc[i,'MIS']))
        select_query="INSERT INTO result VALUES ('" + df_login.loc[i,'NAME'] + "'," + str(df_login.loc[i,'MIS']).strip() + "," + "1)"
        print(select_query)
        # exit(0)
        cursor = conn.cursor()
        cursor.execute(select_query)    

# Close the connection
print('done!!')
conn.commit()
conn.close() 

if __name__ == '__main__':
    app.run(debug=True)
